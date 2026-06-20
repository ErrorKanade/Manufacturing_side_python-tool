import pandas as pd
import os
import csv
from fuzzywuzzy import fuzz
from openpyxl.styles import PatternFill, Alignment, Border, Side, Font as PyFont
from openpyxl.utils import get_column_letter

HAS_RICH_TEXT = False

def print_instructions():
    print("="*75)
    print("           PDCA 测试项批量对比工具(v4.4)")
    print("="*75)
    print(" [使用說明]")
    print(" 1. 請將多個待比對的「子資料夾」放入一個「總資料夾」中。")
    print(" 2. 每個子資料夾內必須包含：")
    print("    - 以 'O' 開頭的舊版 CSV (例如: O_test.csv)")
    print("    - 以 'N' 開頭的新版 CSV (例如: N_test.csv)")
    print(" 3. 將「總資料夾」拖動到此處並按 Enter。")
    print("-" * 75)
    print(" [顏色規則] 黃色: 變更 | 紅色: 新增 | 灰色: 刪除")
    print(" [Kana 2026] ")
    print("="*75)

def read_custom_csv(file_path):
    try:
        with open(file_path, 'r', encoding='utf-8-sig', errors='ignore') as f:
            lines = f.readlines()
        
        # 1. 尋找 "All Parametric Keys" 所在的行號
        start_idx = -1
        for idx, line in enumerate(lines):
            if "All Parametric Keys" in line:
                start_idx = idx
                break
        
        if start_idx == -1:
            print(f"   ⚠️ 找不到關鍵標籤 'All Parametric Keys': {os.path.basename(file_path)}")
            return None

        # 2. 定義相對位置
        def get_line_data(offset):
            target_idx = start_idx + offset
            if target_idx >= len(lines): return []
            # 使用 csv.reader 處理可能有逗號的字串
            reader = csv.reader([lines[target_idx].strip()])
            res = list(reader)
            return res[0] if res else []

        keys = get_line_data(1)        # All Parametric Keys 下一行
        lower_lims = get_line_data(3)  # 下三行
        higher_lims = get_line_data(4) # 下四行

        # 3. 收集資料
        data_list = []
        # 從 index 1 開始
        for i in range(1, len(keys)):
            k = keys[i].strip() if i < len(keys) else ""
            if not k: continue 
            
            low = lower_lims[i].strip() if i < len(lower_lims) else "N/A"
            high = higher_lims[i].strip() if i < len(higher_lims) else "N/A"
            
            data_list.append({'Key': k, 'Lower Lim': low, 'Higher Lim': high})

        df = pd.DataFrame(data_list)
        if df.empty: return None
        
        # 建立唯一 Key 防止重複項造成比對錯誤
        df['Unique_Key'] = df['Key'] + "_" + df.groupby('Key').cumcount().astype(str)
        return df.set_index('Unique_Key')

    except Exception as e:
        print(f"   ⚠️ 讀取出錯 {os.path.basename(file_path)}: {e}")
        return None

def main():
    while True:
        print_instructions()
        parent_input = input("請將「總資料夾」拖動到此處 (或輸入 exit): ").strip().replace("'", "").replace('"', "")
        if parent_input.lower() in ['exit', 'quit', 'q']: break

        parent_path = os.path.abspath(parent_input)
        if not os.path.isdir(parent_path):
            print("❌ 無效路徑！"); continue

        subdirs = [os.path.join(parent_path, d) for d in os.listdir(parent_path) if os.path.isdir(os.path.join(parent_path, d))]
        if not subdirs: subdirs = [parent_path]

        valid_tasks = []
        for folder in subdirs:
            files = os.listdir(folder)
            o = next((f for f in files if f.upper().startswith('O') and f.lower().endswith('.csv')), None)
            n = next((f for f in files if f.upper().startswith('N') and f.lower().endswith('.csv')), None)
            if o and n: valid_tasks.append((folder, o, n))

        if not valid_tasks:
            print("\n⚠️ 找不到 O_*.csv 與 N_*.csv 成對檔案！"); continue

        output_path = os.path.join(parent_path, "所有站位比對彙總表.xlsx")
        summary_data = []

        # 樣式定義
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        yellow_f = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        red_f = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
        gray_f = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        blue_f = PatternFill(start_color='CCEEFF', end_color='CCEEFF', fill_type='solid')
        
        link_font = PyFont(color="0000FF", underline="single", bold=True)
        diff_val_font = PyFont(color="FF0000", bold=True) # 數值變更專用紅字
        normal_font = PyFont(color="000000")

        try:
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                # 建立 Dashboard Sheet
                ws_dash = writer.book.create_sheet('Dashboard彙總', 0)
                
                for folder, o_file, n_file in valid_tasks:
                    folder_name = os.path.basename(folder)
                    sheet_name = folder_name[:31]
                    print(f" 正在處理: {folder_name}...")
                    
                    df_o = read_custom_csv(os.path.join(folder, o_file))
                    df_n = read_custom_csv(os.path.join(folder, n_file))
                    if df_o is None or df_n is None: continue

                    rows = []
                    stats = {"一致": 0, "變更": 0, "新增": 0, "刪除": 0, "名稱微調": 0}
                    o_keys, n_keys = list(df_o.index), list(df_n.index)
                    matched_o, matched_n = set(), set()

                    # 1. 完全匹配與數值比對
                    for ok in o_keys:
                        if ok in n_keys:
                            o_r, n_r = df_o.loc[ok], df_n.loc[ok]
                            o_l, o_h, n_l, n_h = str(o_r['Lower Lim']), str(o_r['Higher Lim']), str(n_r['Lower Lim']), str(n_r['Higher Lim'])
                            
                            is_same = (o_l == n_l and o_h == n_h)
                            st = "完全一致" if is_same else "上下限變更"
                            stats["一致" if is_same else "變更"] += 1
                            nm = ok.rsplit('_', 1)[0]
                            # rows 結構增加舊值參考以便後續標紅判斷
                            rows.append([nm, o_l, o_h, "", nm, n_l, n_h, st, o_l, o_h]) 
                            matched_o.add(ok); matched_n.add(ok)

                    # 2. 模糊匹配
                    unmatched_o = [k for k in o_keys if k not in matched_o]
                    unmatched_n = [k for k in n_keys if k not in matched_n]
                    for ok in unmatched_o:
                        best_nk, max_s = None, 0
                        for nk in unmatched_n:
                            if nk in matched_n: continue
                            s = fuzz.ratio(ok, nk)
                            if s > max_s: max_s = s; best_nk = nk
                        
                        o_nm = ok.rsplit('_', 1)[0]
                        if best_nk and max_s >= 90:
                            n_nm = best_nk.rsplit('_', 1)[0]
                            o_r, n_r = df_o.loc[ok], df_n.loc[best_nk]
                            stats["名稱微調"] += 1
                            rows.append([o_nm, str(o_r['Lower Lim']), str(o_r['Higher Lim']), "", n_nm, str(n_r['Lower Lim']), str(n_r['Higher Lim']), "名稱微調", str(o_r['Lower Lim']), str(o_r['Higher Lim'])])
                            matched_n.add(best_nk); matched_o.add(ok)
                        else:
                            stats["刪除"] += 1
                            rows.append([o_nm, str(df_o.loc[ok]['Lower Lim']), str(df_o.loc[ok]['Higher Lim']), "", "", "", "", "項目刪除", "", ""])

                    # 3. 新增
                    for nk in unmatched_n:
                        if nk not in matched_n:
                            stats["新增"] += 1
                            rows.append(["", "", "", "", nk.rsplit('_', 1)[0], str(df_n.loc[nk]['Lower Lim']), str(df_n.loc[nk]['Higher Lim']), "新增項目", "", ""])

                    # --- 寫入分頁 ---
                    summary_data.append({'站位': folder_name, 'Sheet': sheet_name, **stats})
                    ws = writer.book.create_sheet(sheet_name)
                    
                    # 1. 第一行：返回連結
                    back_cell = ws.cell(row=1, column=1, value="🔙 返回彙總表")
                    back_cell.hyperlink = f"#'Dashboard彙總'!A1"
                    back_cell.font = link_font
                    
                    # 2. 第二行：寫入源文件名稱 
                    # 舊版文件名 (跨 A-C 欄)
                    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=3)
                    o_title = ws.cell(row=2, column=1, value=f"源文件(舊): {o_file}")
                    o_title.font = PyFont(bold=True, color="4472C4") # 藍色
                    o_title.alignment = Alignment(horizontal='center')

                    # 新版文件名 (跨 E-G 欄)
                    ws.merge_cells(start_row=2, start_column=5, end_row=2, end_column=7)
                    n_title = ws.cell(row=2, column=5, value=f"源文件(新): {n_file}")
                    n_title.font = PyFont(bold=True, color="ED7D31") # 橘色
                    n_title.alignment = Alignment(horizontal='center')

                    # 3. 第三行：寫入藍色表頭 (這裡 row 必須是 3)
                    headers = ["項目名稱 (舊)", "下限", "上限", "VS", "項目名稱 (新)", "下限", "上限", "比對狀態"]
                    for c, h in enumerate(headers, 1):
                        cell = ws.cell(row=3, column=c, value=h)
                        cell.fill = blue_f
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal='center')

                    # 4. 第四行起：寫入數據 (注意 enumerate 的起始值是 4)                            
                    for r_idx, r_data in enumerate(rows, 4):
                        status = r_data[7]  # 取得比對狀態
                        
                        # 從 r_data 中提取預留的舊值參考 (用於比對數值是否變更)
                        old_l_ref = r_data[8] 
                        old_h_ref = r_data[9]
                        
                        for c_idx, val in enumerate(r_data[:8], 1):
                            cell = ws.cell(row=r_idx, column=c_idx, value=val)
                            cell.border = thin_border
                            
                            # 顏色與字體邏輯
                            if status == "上下限變更":
                                cell.fill = yellow_f
                                # 如果是新下限(6)且不等於舊下限，或是新上限(7)且不等於舊上限，標紅
                                if (c_idx == 6 and str(val) != str(old_l_ref)) or \
                                   (c_idx == 7 and str(val) != str(old_h_ref)):
                                    cell.font = diff_val_font
                            elif status == "新增項目":
                                cell.fill = red_f
                            elif status == "項目刪除":
                                cell.fill = gray_f
                            elif status == "名稱微調":
                                cell.fill = yellow_f

                    # 5. 自動調整欄寬
                    for col in range(1, 9):
                        ws.column_dimensions[get_column_letter(col)].width = 25

                # --- 寫入 Dashboard 彙總頁 ---
                df_summary = pd.DataFrame(summary_data)
                df_summary.to_excel(writer, sheet_name='Dashboard彙總', index=False)
                ws_dash = writer.sheets['Dashboard彙總']
                # 為 Dashboard 增加超連結指向各分頁
                for r_idx, row in enumerate(summary_data, 2):
                    cell = ws_dash.cell(row=r_idx, column=2) # Sheet 欄位
                    cell.hyperlink = f"#'{row['Sheet']}'!A1"
                    cell.font = link_font

            print(f"\n✅ 處理完成！檔案儲存於: {output_path}")

        except Exception as e:
            print(f"❌ 發生錯誤: {e}")

if __name__ == "__main__":
    main()